import requests
import io
from openpyxl import load_workbook
import json
from collections import defaultdict, Counter
from datetime import datetime, timedelta
from openpyxl.styles import PatternFill

def find_group_column(sheet, group_name):
    for row in sheet.iter_rows(min_row=1, max_row=24):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and group_name in cell.value:
                return cell.column
    return None

def split_time_interval(time_str):
    try:
        time_str = time_str.replace("–", "-").replace("—", "-").strip()
        parts = time_str.split("-")
        if len(parts) != 2:
            return time_str, time_str

        start_str, end_str = parts
        start_str = start_str.replace(".", ":").strip()
        end_str = end_str.replace(".", ":").strip()

        start = datetime.strptime(start_str, "%H:%M")
        end = datetime.strptime(end_str, "%H:%M")
        total_minutes = int((end - start).total_seconds() // 60)

        half = (total_minutes - 10) // 2
        first_end = start + timedelta(minutes=half)
        second_start = first_end + timedelta(minutes=10)

        first_interval = f"{start.strftime('%H:%M')} - {first_end.strftime('%H:%M')}"
        second_interval = f"{second_start.strftime('%H:%M')} - {end.strftime('%H:%M')}"
        return first_interval, second_interval
    except:
        return time_str, time_str

def interpret_color(cell):
    try:
        fill = cell.fill
        fg = fill.fgColor
        if fg.type == "rgb":
            rgb = fg.rgb.upper()
            if rgb.startswith("FF"):
                rgb = rgb[2:]

            upper_colors = {"00B0F0", "0070C0", "0066FF", "CCECFF"}
            lower_colors = {"00FF00", "00B050", "00FF99"}
            hour_colors = {"FFC0CB", "FF99CC", "FFB6C1", "FF66CC"}

            if rgb in upper_colors:
                return "upper"
            elif rgb in lower_colors:
                return "lower"
            elif rgb in hour_colors:
                return "hour"
        return None
    except:
        return None

def get_excel_schedule(url, group_name):
    try:
        response = requests.get(url)
        response.raise_for_status()

        excel_file_in_memory = io.BytesIO(response.content)
        workbook = load_workbook(excel_file_in_memory)
        sheet = workbook.active
        
        column_index = find_group_column(sheet, group_name)
        if not column_index:
            print(f"Ошибка: Не удалось найти столбец для группы {group_name}.")
            return None

        schedule_data = {
            "group": group_name,
            "schedule": []
        }
        
        current_day = None
        current_time = None
        pair_counts = defaultdict(int)

        for row in sheet.iter_rows(min_row=3): 
            row_values = [cell.value for cell in row]

            day_cell_value = row_values[1] if len(row_values) > 1 else None
            time_cell_value = row_values[2] if len(row_values) > 2 else None

            if day_cell_value:
                current_day = str(day_cell_value).strip()
            if time_cell_value:
                current_time = str(time_cell_value).strip()

            pair_number = None
            clean_time = current_time
            parts = current_time.split() if current_time else []
            if parts and parts[0].isdigit():
                pair_number = parts[0]
                clean_time = " ".join(parts[1:]) if len(parts) > 1 else current_time

            subject_cell = row_values[column_index] if column_index < len(row_values) else None
            room_cell = row_values[column_index + 1] if column_index + 1 < len(row_values) else None

            if subject_cell:
                key = (current_day, pair_number)
                pair_counts[key] += 1

                color_type = interpret_color(row[column_index])
                week_type = None
                duration = 2

                if color_type == "upper":
                    week_type = "upper"
                elif color_type == "lower":
                    week_type = "lower"
                elif color_type == "hour":
                    duration = 1

                indexed_pair = f"{pair_number}/{pair_counts[key]}" if pair_number else None

                pair_info = {
                    "day": current_day,
                    "time": clean_time,
                    "raw_time": clean_time,
                    "pair": indexed_pair,
                    "room": str(room_cell).strip() if room_cell else "",
                    "subject": str(subject_cell).strip() if subject_cell else "",
                    "week_type": week_type,
                    "duration": duration
                }

                schedule_data["schedule"].append(pair_info)

        pair_occurrences = Counter((p["day"], p["pair"].split("/")[0]) for p in schedule_data["schedule"] if p.get("pair"))
        interval_cache = {}

        for entry in schedule_data["schedule"]:
            pair_raw = entry.get("pair")
            if not pair_raw or "/" not in pair_raw:
                continue

            base_pair = pair_raw.split("/")[0]
            index = int(pair_raw.split("/")[1])
            key = (entry["day"], base_pair)

            if key not in interval_cache:
                interval_cache[key] = split_time_interval(entry["time"])

            first, second = interval_cache[key]
            entry["time"] = first if index == 1 else second

            if pair_occurrences[key] == 1:
                entry["pair"] = f"{base_pair}/1" if entry.get("duration") == 1 else base_pair

        return schedule_data
        
    except requests.exceptions.RequestException as e:
        print(f"Ошибка при скачивании файла: {e}")
        return None
    except Exception as e:
        print(f"Ошибка при обработке файла Excel: {e}")
        return None

if __name__ == '__main__':
    EXCEL_URL = "http://www.bobruisk.belstu.by/uploads/b1/s/8/648/basic/117/614/Raspisanie_uchebnyih_zanyatiy_na_2025-2026_uch.god_1_semestr.xlsx?t=1756801696"
    MY_GROUP = "РС02-24"
    
    my_schedule = get_excel_schedule(EXCEL_URL, MY_GROUP)
    
    if my_schedule:
        print("Получено расписание:")
        print(json.dumps(my_schedule, indent=4, ensure_ascii=False))
